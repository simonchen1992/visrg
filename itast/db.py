import MySQLdb
from sys import stdout

# Utility for printing in colourful
from colorama import init
#from colorama import Fore, Back, Style


init()

# # Return values, and either to be considered PASS or FAIL
# APPROVED_OFFLINE = '5931'
# APPROVED_ONLINE = '3030'
# DECLINED_OFFLINE = '5A31'
# INCOMPLETED = '0000'
# TERMINATED_ERROR_SW = 'EF01'
# TERMINATED_INCORRECT_DATA = 'EF02'
# TERMINATED_TIMEOUT = 'EF03'
# TERMINATED_COLLISION = 'EF04'
# TERMINATED_REQ_RESET = 'EF05'
# TERMINATED_SWITCH_IFACE = 'EF06'
# TERMINATED_REF_MOBILE = 'EF07'
#
# TX_PASS = (APPROVED_ONLINE, APPROVED_OFFLINE)
# TX_DEC = (DECLINED_OFFLINE)
# TX_FAIL = (DECLINED_OFFLINE,
#            INCOMPLETED,
#            TERMINATED_ERROR_SW,
#            TERMINATED_INCORRECT_DATA,
#            TERMINATED_TIMEOUT,
#            TERMINATED_COLLISION,
#            TERMINATED_REQ_RESET,
#            TERMINATED_SWITCH_IFACE,
#            TERMINATED_REF_MOBILE)
# TX_TF = (  TERMINATED_ERROR_SW,
#            TERMINATED_INCORRECT_DATA,
#            TERMINATED_SWITCH_IFACE)
# TX_CF = (  TERMINATED_TIMEOUT,
#            TERMINATED_COLLISION,
#            TERMINATED_REQ_RESET)
# TX_UC = (INCOMPLETED)


def connect():
    db = MySQLdb.connect(user='itast',
                       passwd='APtsd01$',
                       host='192.168.48.195',
                       db='itast')
    return db
#
# def print_table(fields, cursor):
#   for f in fields:
#     stdout.write(Style.BRIGHT + f.ljust(fields[f]+2).capitalize()[:fields[f]+2] + Style.RESET_ALL)
#   print ""
#   for row in cursor.fetchallDict():
#     for f in fields:
#       stdout.write(str(row[f]).strip().ljust(fields[f]+2)[:fields[f]+2])
#     print ""
#
# def print_sessions(db, keyword):
#   fields = {}
#   fields['id'] = 8
#   fields['created'] = 20
#   fields['expedient'] = 32
#   fields['owner'] = 16
#   fields['dut1_name'] = 16
#   fields['dut1_description'] = 24
#   query_fileds = ','.join(fields.keys())
#   cur = db.cursor(cursorclass=MySQLdb.cursors.DictCursor)
#   cur.execute("SELECT " + query_fileds + " FROM test_sessions " +
#               "WHERE expedient LIKE '%" + keyword + "%' " +
#               "OR id LIKE '%" + keyword + "%' " +
#               "OR dut1_id LIKE '%" + keyword + "%' " +
#               "OR dut1_name LIKE '%" + keyword + "%' " +
#               "OR dut1_description LIKE '%" + keyword + "%' " +
#               "OR dut2_id LIKE '%" + keyword + "%' " +
#               "OR dut2_name LIKE '%" + keyword + "%' " +
#               "OR dut2_description LIKE '%" + keyword + "%'" +
#               "ORDER BY created DESC LIMIT 10;")
#   print_table(fields, cur)
#
# def set_position_verdict(cardpostx):
#   txPass = 0
#   txFail = 0
#   txCF = 0
#   txTF = 0
#   txUC = 0
#   txDEC = 0
#   txTotal = 0
#   for tx in cardpostx['txs']:
#     txTotal =+ txTotal + 1
#     if cardpostx['txs'][tx] in TX_PASS:
#       txPass = txPass + 1
#     #elif cardpostx['txs'][tx] in TX_FAIL:
#     #  txFail =+ txFail + 1
#     elif cardpostx['txs'][tx] in TX_CF:
#       txCF = txCF + 1
#     elif cardpostx['txs'][tx] in TX_TF:
#       txTF = txTF + 1
#     elif cardpostx['txs'][tx] in TX_UC:
#       txUC = txUC + 1
#     else:
#       print "What the hell is this result: " + cardpostx['txs'][tx]
#
#   if txPass >= 5:
#     cardpostx['verdict'] = "OK"
#   elif txCF + txTF > 0 and txPass > 0:
#     cardpostx['verdict'] = "DF"
#   elif txCF >= 3:
#     cardpostx['verdict'] = "CF"
#   elif txCF < 3 and txCF + txTF >=3:
#     cardpostx['verdict'] = "TF"
#   elif txTotal == 0:
#     cardpostx['verdict'] = "  "
#   else:
#     cardpostx['verdict'] = "KO"

def load_session_data(db, idsession, txlimit=7):
    """
    Get session data from DB
    """
    session = {}
    cur = db.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    cur.execute("SELECT * FROM test_sessions WHERE id = " + idsession + ";")
    session = cur.fetchoneDict()
    return session

def load_card_results_by_session(db, idsession, txlimit=1):
    """
		Get card data and test case data from DB
    """
    print('load database start\n')
    visaCards = {}
    cur = db.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    # Get Visa Cards from DB
    cur.execute("SELECT * FROM visa_cards WHERE active = 1 ORDER BY vtf ASC;")
    visaCards = cur.fetchallDict()
    
    # For each card in Visa card deck, get the most recent 10 txs from DB
    for card in visaCards:
        card['txs'] = {}
        card['threefail_flag'] = False
        for pos in getPositions():
            card['txs'][pos] = {}
            card['txs'][pos]['verdict'] = ""
            card['txs'][pos]['comment'] = ""
            # todo: DUT is set to 1 always, configure later
            cur.execute("SELECT verdict,comments FROM test_cases " +
                      "WHERE id_test_session = " + str(idsession) + " AND dut = '1' AND id_card = " + str(card['id']) + " AND pos = '" + pos + "' " +
                      "ORDER BY created DESC LIMIT 1")
            for row in cur.fetchallDict():
                # check if there's any fail in z=3 position
                if pos in ['3C','3N','3S','3E','3W'] and row['verdict'] in ['CF', 'TF', 'DF']:
                    card['threefail_flag'] = True
                card['txs'][pos]['verdict'] = row['verdict']
                card['txs'][pos]['comment'] = row['comments']
    print('load database finished\n')
    return visaCards

def export_results_to_excel(cards, wb, txlimit=1):
    from openpyxl.styles import colors, PatternFill
    style_red = PatternFill(fill_type='solid', fgColor=colors.RED)
    style_green = PatternFill(fill_type='solid', fgColor='00008000')
    style_yellow = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
    style_white = PatternFill(fill_type='solid', fgColor=colors.WHITE)
    writeReportFlag = True  # bool writeReportFlag: indicate if information is enough to write on "VISA REPORT"
    cardsData = {}  # {dictionary} cardsData: store information used for "VISA REPORT"
    """
        write position verdict from database into ROADMAP
    """
    wsRoadmap = wb.create_sheet(title="Roadmap")
    wsRoadmap.append(['Card VTF','ID','Verdict'] + getPositions() + ['Comment'])  # add title row
    for rowID, card in enumerate(cards):
        rowID += 2  # fit the row in the excel
        cardData = []  # [list] cardData: store information for single row of "ROADMAP"
        cardsData[card['vtf'].strip()] = {}
        cardData.append(card['vtf'].strip())
        cardData.append(card['id'])
        cardData.append('XX')
        # determine required test position for each card
        ntpos = testposition(card['positions'])
        cardsData[card['vtf'].strip()]['positions'] = card['positions']
        if card['threefail_flag']:
            cardsData[card['vtf'].strip()]['positions'] = card['positions'].replace(' NT@2N, NT@2S, NT@2E, NT@2W,', '')
            ntpos.remove('2W')
            ntpos.remove('2E')
            ntpos.remove('2N')
            ntpos.remove('2S')
        # cardData only collect those test results for required test positions
        for pos in getPositions():
            if pos in ntpos:
                cardData.append('')
            elif str(card['txs'][pos]['verdict']) in ['', 'Pending']:
                cardData.append('Pending')
            else:
                cardData.append(str(card['txs'][pos]['verdict']))
        # update cardData results for TF and CF;
        # rules 1: if there appears FAIL and PASS position verdict for one card, all FAIL verdict are determined as DF
        # rules 2: if there appears only FAIL verdict for one card, but it contains both CF and TF, all FAIL verdict are determined as TF
        if 'CF' in cardData or 'TF' in cardData:
            if 'P' in cardData:
                cardData = ['DF' if i in ['CF','TF'] else i for i in cardData]
            elif 'CF' in cardData and 'TF' in cardData:
                cardData = ['TF' if i == 'CF' else i for i in cardData]
        cardData.append(str(card['txs']['0N']['comment']))  # add comments for one card
        wsRoadmap.append(cardData)
        """
            determine card verdict in ROADMAP
        """
        determine_flag = False
        null_flag = True
        failPosition = ''  # store fail positions for "VISA REPORT"
        for posVerdict in wsRoadmap[str(rowID)]:
            if posVerdict.value == 'Pending':
                wsRoadmap[str(posVerdict.column) + str(rowID)].fill = style_yellow
                wsRoadmap['C' + str(rowID)].value = 'Pending'
                wsRoadmap['C' + str(rowID)].fill = style_yellow
                determine_flag = True
                writeReportFlag = False
            elif posVerdict.value in ['CF', 'TF', 'DF']:
                wsRoadmap[str(posVerdict.column) + str(rowID)].fill = style_red
                null_flag = False
                if not determine_flag:
                    wsRoadmap['C' + str(rowID)].value = 'F'
                    wsRoadmap['C' + str(rowID)].fill = style_red
                    determine_flag = True
                # write to visa template
                failPosition += posVerdict.value + '@' + wsRoadmap[str(posVerdict.column) + '1'].value + ', '  # for example: DF@2N,
            elif posVerdict.value == 'P':
                wsRoadmap[str(posVerdict.column) + str(rowID)].fill = style_green
                null_flag = False
        if null_flag:
            wsRoadmap['C' + str(rowID)].value = ''
            wsRoadmap['C' + str(rowID)].fill = style_white
        if not null_flag and not determine_flag:
            wsRoadmap['C' + str(rowID)].value = 'P'
            wsRoadmap['C' + str(rowID)].fill = style_green
        """
            store results and position information for VISA REPORT
        """
        if wsRoadmap['C' + str(rowID)].value == 'P':
            cardsData[card['vtf'].strip()]['result'] = 'Pass'
        elif wsRoadmap['C' + str(rowID)].value == 'F':
            cardsData[card['vtf'].strip()]['result'] = 'Fail'
        else:
            writeReportFlag = False
        # replace for failPosition when all position fail for height, e.g. DF@3A
        for failType in ['DF', 'CF', 'TF']:
            for height in ['0', '1', '2', '3']:
                failPosition = failPosition.replace('%s@%sN, %s@%sS, %s@%sE, %s@%sW, %s@%sC, ' % (failType, height, failType, height, failType, height, failType, height, failType, height),
                                                    '%s@%sA, ' % (failType, height))
        cardsData[card['vtf'].strip()]['positions'] = failPosition + cardsData[card['vtf'].strip()]['positions']
    
    """
        write cardsData into VISA REPORT if writeReportFlag is true
    """
    # write to visa result template
    # if not writeReportFlag:
    #     print 'The report is not complete, please check roadmap!\n'
    #     return
    wsVisaReport = wb['Cross_test_results']
    for row in wsVisaReport.iter_rows(min_row=4, min_col=2, max_col=2):
        cardVtf = row[0]
        if cardVtf.value in [None, '']:
            break
        try:
            resultCol = 'D'
            locationCol = 'E'
            positionCol = 'G'
            wsVisaReport[resultCol + str(cardVtf.row)].value = cardsData[cardVtf.value]['result']
            wsVisaReport[locationCol + str(cardVtf.row)].value = 'Applus Shanghai'
            wsVisaReport[positionCol + str(cardVtf.row)].value = cardsData[cardVtf.value]['positions']
        except KeyError as e:
            print('Could not find vtf in VISA REPORT, please check the version of VISA template')
            raise KeyError(str(e))


def testposition(st):
    ntpos = str(st).replace('NT@', '')
    ntpos = ntpos.replace(' ', '')
    ntpos = ntpos.replace('0A', '0C,0N,0S,0E,0W')
    ntpos = ntpos.replace('1A', '1C,1N,1S,1E,1W')
    ntpos = ntpos.replace('2A', '2C,2N,2S,2E,2W')
    ntpos = ntpos.replace('3A', '3C,3N,3S,3E,3W')
    ntpos = ntpos.replace('4A', '4C,4N,4S,4E,4W')
    ntpos = ntpos.split(',')
    #testpos = [i for i in getPositions() if i not in ntpos]
    return ntpos


def getPositions():
    positions = []
    for Z in ['0','1','2','3','4']: # Iterate Z
        for XY in ['N', 'S', 'E', 'W', 'C']: # Iterate cardinals
            positions.append(Z+XY)
    return positions

def print_xtroadmap(cards):
    for card in cards:
        stdout.write(card['vtf'].strip() + "\t")
        for pos in getPositions():
            stdout.write(str(card['txs'][pos]['verdict']) + " ")
        print ""












